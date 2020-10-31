import openpyxl, random
from openpyxl import load_workbook
wb = load_workbook('./tablet.xlsx')
worksheet1 = wb['Взносы']
worksheet1['A1'] = 'Фамилия'
worksheet1['b1'] = 'Взнос'
worksheet1['c1'] = 'Месяц'

all_camps = ['Бухта','Лесной','Солнечный']
all_months = ['Сентябрь','Октябрь','Ноябрь','Декабрь','Январь','Февраль','Март','Апрель','Май']
all_surnames = ['Смирнов', 'Иванов', 'Кузнецов', 'Соколов', 'Попов', 'Лебедев', 'Козлов', 'Новиков', 'Морозов', 'Петров', 'Волков', 'Соловьёв', 'Васильев']

for a in range(2,len(all_surnames)+2):
    worksheet1.cell(row=a, column=1).value = all_surnames[a-2];

    b = random.randint(0,100)
    if b >= 25:
        worksheet1.cell(row=a, column=2).value = random.randint(45,70)
    else:
        worksheet1.cell(row=a, column=2).value = 0
    if worksheet1.cell(row=a, column=2).value > 0:
        worksheet1.cell(row=a, column=3).value = random.choice(all_months)
    else:
        worksheet1.cell(row=a, column=3).value = ' - '

# Первая страница таблицы заполнена

worksheet2 = wb['Путёвки']
worksheet2['a1'] = 'Фамилия'
worksheet2['b1'] = 'Путевка'
worksheet2['c1'] = 'Лагерь'

for c in range(2,len(all_surnames)+2):
    worksheet2.cell(row=c, column=1).value = worksheet1.cell(row=c, column=1).value
    if worksheet1.cell(row=c, column=2).value>0:
        worksheet2.cell(row=c, column=2).value = 'Получена'
    else:
        worksheet2.cell(row=c, column=2).value = ' - '
    if worksheet2.cell(row=c, column=2).value ==' - ':
        worksheet2.cell(row=c,column=3).value = ' - '
    else:
        worksheet2.cell(row=c, column=3).value = random.choice(all_camps)

# Вторая страница таблицы заполнена

wb.save('./tablet.xlsx')